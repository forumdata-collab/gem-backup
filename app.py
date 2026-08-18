#!/usr/bin/env python3
"""Gem Backup 網頁 — 用戶俾 public Drive gem folder URL → xlsx (+zip), 30 分鐘自動刪."""
import json, os, re, threading, time, uuid, zipfile
from dataclasses import dataclass
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE = '/tmp/gembackup'
TTL = 30 * 60            # 檔案保留 30 分鐘
CELL_MAX = 30000         # Excel cell 上限 32767，留 buffer
MAX_FILE = 200 * 1024 * 1024    # 單一附件上限
MAX_TOTAL = 1024 * 1024 * 1024  # 每 job 總量上限（有 enforce）
MAX_ACTIVE = 4
MAX_DEPTH = 2            # 子 folder 遞歸深度

FOLDER_RE = re.compile(r'drive\.google\.com/drive/(?:u/\d+/)?folders/([A-Za-z0-9_-]+)')
ENTRY_RE = re.compile(r'flip-entry" id="entry-([A-Za-z0-9_-]+)".*?'
                      r'href="(https://[^"]+)".*?'
                      r'flip-entry-title">([^<]+)</div>.*?'
                      r'flip-entry-last-modified"><div>([^<]*)</div>', re.S)

_lock = threading.Lock()
_active = {}


@dataclass
class Entry:
    fid: str
    name: str
    mtime: str
    kind: str  # 'gem' | 'file' | 'folder'


@dataclass
class GemRow:
    name: str
    desc: str
    instr: str
    fid: str
    mtime: str


def read_varint(b, i):
    r = 0; s = 0
    while True:
        x = b[i]; i += 1
        r |= (x & 0x7f) << s
        if not x & 0x80: return r, i
        s += 7


def walk_fields(b):
    i = 0; out = []
    while i < len(b):
        try: tag, i = read_varint(b, i)
        except IndexError: break
        fnum, wt = tag >> 3, tag & 7
        if wt == 0:
            try: v, i = read_varint(b, i)
            except IndexError: break
            out.append((fnum, 'v', v))
        elif wt == 2:
            ln, i = read_varint(b, i)
            out.append((fnum, 'b', b[i:i+ln])); i += ln
        else:
            break
    return out


def get_str(fields, n):
    for fn, wt, v in fields:
        if fn == n and wt == 'b':
            return v.decode('utf-8', 'replace')
    return ''


def parse_gem(b):
    rows = []
    for fn, wt, v in walk_fields(b):
        if fn == 2 and wt == 'b':
            sub = walk_fields(v)
            name, desc, instr = get_str(sub, 1), get_str(sub, 2), get_str(sub, 3)
            if name or instr:
                rows.append((name, desc, instr))
    return rows


def sanitize(name):
    return re.sub(r'[^\w.\-]+', '_', name) or 'file'


def dedupe_name(base, used):
    used[base] = used.get(base, 0) + 1
    if used[base] == 1:
        return base
    stem, dot, ext = base.rpartition('.')
    if dot and len(ext) <= 5:
        return f'{stem} ({used[base]-1}).{ext}'
    return f'{base}_{used[base]-1}'


def _stream(r, dest, cap):
    total = 0
    with open(dest, 'wb') as f:
        for chunk in r.iter_content(1 << 16):
            total += len(chunk)
            if total > cap: raise RuntimeError('file too large')
            f.write(chunk)
    if total == 0: raise RuntimeError('empty download')


def dl_to_file(fid, dest, cap=MAX_FILE):
    url = f'https://drive.google.com/uc?export=download&id={fid}'
    with requests.get(url, stream=True, timeout=60) as r:
        if r.headers.get('Content-Type', '').startswith('text/html') and 'confirm' in r.text:
            m = re.search(r'confirm=([0-9A-Za-z_-]+)', r.text)
            if not m: raise RuntimeError('confirm token not found')
            with requests.get(url, params={'confirm': m.group(1)}, stream=True, timeout=120) as r2:
                _stream(r2, dest, cap)
        else:
            _stream(r, dest, cap)


def classify(href):
    if '/gem/' in href: return 'gem'
    if '/file/d/' in href: return 'file'
    if '/folders/' in href: return 'folder'
    return None


def list_folder(folder_id, depth=0, seen=None):
    if depth > MAX_DEPTH:
        return []
    seen = seen if seen is not None else set()
    if folder_id in seen:
        return []
    seen.add(folder_id)
    r = requests.get(f'https://drive.google.com/embeddedfolderview?id={folder_id}#list', timeout=30)
    r.raise_for_status()
    out = []
    for fid, href, name, mtime in ENTRY_RE.findall(r.text):
        kind = classify(href)
        if kind is None:
            continue
        if kind == 'folder':
            out.extend(list_folder(fid, depth + 1, seen))
        else:
            out.append(Entry(fid, name.strip(), mtime.strip(), kind))
    return out


def split_cell(s):
    if len(s) <= CELL_MAX:
        return [s]
    return [s[i:i + CELL_MAX] for i in range(0, len(s), CELL_MAX)]


def make_xlsx(rows, dest):
    wb = Workbook(); ws = wb.active; ws.title = 'Gems'
    headers = ['名稱', '描述 (Description)', '指示 (Instructions / System Prompt)', 'Gem 連結', '最後修改']
    ws.append(headers)
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c); cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='top')
    for row in rows:
        descs, instrs = split_cell(row.desc), split_cell(row.instr)
        n = max(len(descs), len(instrs))
        for i in range(n):
            ws.append([
                row.name if i == 0 else '',
                descs[i] if i < len(descs) else '',
                ('【續上】' + instrs[i]) if i > 0 and i < len(instrs) else (instrs[i] if i < len(instrs) else ''),
                f'https://gemini.google.com/gem/{row.fid}' if i == 0 else '',
                row.mtime if i == 0 else '',
            ])
    for i, w in enumerate([24, 60, 90, 45, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    wb.save(dest)


def set_state(job_id, **kw):
    st = {'created': time.time(), 'files': [], **kw}
    json.dump(st, open(os.path.join(BASE, job_id, 'state.json'), 'w'))


def build_package(job_id, folder_id):
    jd = os.path.join(BASE, job_id)
    entries = list_folder(folder_id)
    if not entries:
        raise RuntimeError('Folder 未公開或不存在 — 請喺 Drive 將 folder 設為「Anyone with the link」')
    gems = [e for e in entries if e.kind == 'gem']
    atts = [e for e in entries if e.kind == 'file']
    total = 0
    rows = []
    for e in gems:
        dest = os.path.join(jd, 'raw_' + e.fid)
        dl_to_file(e.fid, dest)
        total += os.path.getsize(dest)
        if total > MAX_TOTAL: raise RuntimeError('檔案總量超過 1GB 上限')
        parsed = parse_gem(open(dest, 'rb').read()) or [(e.name, '', '')]
        os.unlink(dest)
        rows += [GemRow(n or e.name, d, i, e.fid, e.mtime) for n, d, i in parsed]
    xlsx_path = os.path.join(jd, 'Gems_Backup.xlsx')
    make_xlsx(rows, xlsx_path)
    files = [{'name': 'Gems_Backup.xlsx', 'size': os.path.getsize(xlsx_path)}]
    if atts:
        zip_path = os.path.join(jd, 'Gems_Attachments.zip')
        used = {}
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for e in atts:
                dest = os.path.join(jd, 'att_' + e.fid)
                try:
                    dl_to_file(e.fid, dest)
                    total += os.path.getsize(dest)
                    if total > MAX_TOTAL: raise RuntimeError('檔案總量超過 1GB 上限')
                    z.write(dest, dedupe_name(sanitize(e.name), used))
                    os.unlink(dest)
                except RuntimeError:
                    raise
                except Exception:
                    try: os.unlink(dest)
                    except OSError: pass
        files.append({'name': 'Gems_Attachments.zip', 'size': os.path.getsize(zip_path)})
    set_state(job_id, state='done', files=files, summary=f'{len(rows)} 個 Gem × {len(atts)} 個附件')


def run_job(job_id, folder_id):
    try:
        os.makedirs(os.path.join(BASE, job_id), exist_ok=True)
        build_package(job_id, folder_id)
    except Exception as e:
        set_state(job_id, state='error', error=str(e))
    finally:
        with _lock:
            _active.pop(job_id, None)


def sweep():
    now = time.time()
    if not os.path.isdir(BASE): return
    for name in os.listdir(BASE):
        p = os.path.join(BASE, name)
        try:
            if now - os.path.getmtime(p) > TTL:
                import shutil; shutil.rmtree(p, ignore_errors=True)
        except OSError:
            pass


def job_state(job_id):
    if not re.fullmatch(r'[0-9a-f]{32}', job_id or ''):
        return None
    p = os.path.join(BASE, job_id, 'state.json')
    if not os.path.exists(p):
        return None
    try: return json.load(open(p))
    except Exception: return None


PAGE = """<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Gem Backup</title>
<style>
body{margin:0;background:#0f1115;color:#e6e6e6;font-family:-apple-system,'Segoe UI',Roboto,'Noto Sans TC',sans-serif;min-height:100vh}
.wrap{max-width:680px;margin:0 auto;padding:48px 20px}
h1{font-size:26px;font-weight:600;margin:0 0 6px}
p.sub{color:#8b93a7;margin:0 0 28px;font-size:14px;line-height:1.6}
input{width:100%;box-sizing:border-box;padding:14px 16px;border-radius:10px;border:1px solid #2a3040;background:#161a22;color:#fff;font-size:15px;outline:none}
input:focus{border-color:#4a6cf7}
button{margin-top:14px;width:100%;padding:14px;border:0;border-radius:10px;background:#4a6cf7;color:#fff;font-size:15px;font-weight:600;cursor:pointer}
button:disabled{opacity:.5;cursor:default}
#status{margin-top:22px;font-size:14px;line-height:1.7;display:none;background:#161a22;border:1px solid #2a3040;border-radius:10px;padding:16px}
#status.err{border-color:#7a3030;color:#ff9c9c}
#status.done{border-color:#2a5c38}
a.dl{display:inline-block;margin:8px 8px 0 0;padding:10px 18px;border-radius:8px;background:#1c5630;color:#7dffa9;text-decoration:none;font-weight:600}
.note{color:#5c6475;font-size:12px;margin-top:20px;line-height:1.7}
</style></head><body><div class="wrap">
<h1>Gem Backup</h1>
<p class="sub">貼上你嘅 Google Drive Gemini Gem folder 連結（folder 要設為「Anyone with the link」），伺服器會將所有 Gem 整合成 Excel (.xlsx)，附件一併打包成 .zip 俾你下載。檔案只保留 30 分鐘。</p>
<input id="url" placeholder="https://drive.google.com/drive/folders/XXXX" autocomplete="off">
<button id="go">開始 Backup</button>
<div id="status"></div>
<p class="note">Gem 係 Gemini 嘅自訂指令。2026 年 10 月起 Google 停止 Gems 功能，呢個工具幫你將 Gem 嘅指令完整保留下嚟。<br>伺服器唔儲存任何內容 — 30 分鐘後自動刪除。</p>
</div>
<script>
const $=id=>document.getElementById(id);
$('go').onclick=async()=>{
  const url=$('url').value.trim();
  if(!url)return;
  const st=$('status'); st.className=''; st.style.display='block'; st.textContent='建立工作…';
  $('go').disabled=true;
  try{
    const r=await fetch('/api/backup',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({url})});
    const d=await r.json();
    if(d.error){st.className='err';st.textContent='錯誤：'+d.error;return;}
    st.textContent='處理緊：連結 folder → 下載 Gem → 分析指令 → 打包…（大附件可能要一兩分鐘）';
    const t=setInterval(async()=>{
      const s=await (await fetch('/api/status?job='+d.job)).json();
      if(s.state==='done'){
        clearInterval(t);
        st.className='done';
        st.innerHTML='✅ '+s.summary+'<br>下載 (30 分鐘內有效)：';
        s.files.forEach(f=>{st.insertAdjacentHTML('beforeend',`<a class="dl" href="/d/${d.job}/${encodeURIComponent(f.name)}">${f.name} (${(f.size/1048576).toFixed(1)}MB)</a>`)});
        $('go').disabled=false;
      }else if(s.state==='error'){
        clearInterval(t); st.className='err';
        st.textContent='錯誤：'+s.error; $('go').disabled=false;
      }else st.textContent='處理緊：連結 folder → 下載 Gem → 分析指令 → 打包…（大附件可能要一兩分鐘）';
    },2000);
  }catch(e){st.className='err';st.textContent='錯誤：'+e.message;$('go').disabled=false;}
};
</script></body></html>"""


class H(BaseHTTPRequestHandler):
    def log_message(self, *a): pass
    def _json(self, obj, code=200):
        b = json.dumps(obj, ensure_ascii=False).encode()
        self.send_response(code); self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(b))); self.end_headers(); self.wfile.write(b)
    def do_GET(self):
        sweep()
        p = urlparse(self.path)
        if p.path == '/':
            b = PAGE.encode(); self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', str(len(b))); self.end_headers(); self.wfile.write(b)
        elif p.path == '/api/status':
            job = parse_qs(p.query).get('job', [''])[0]
            self._json(job_state(job) or {'state': 'pending'})
        elif p.path.startswith('/d/'):
            parts = p.path.split('/')[2:]
            if len(parts) != 2: self._json({'error': 'bad path'}, 404); return
            job, name = parts
            s = job_state(job)
            if not s or s.get('state') != 'done': self._json({'error': 'not ready'}, 404); return
            fp = os.path.realpath(os.path.join(BASE, job, name))
            if not fp.startswith(os.path.realpath(os.path.join(BASE, job)) + os.sep) or not os.path.isfile(fp):
                self._json({'error': 'not found'}, 404); return
            size = os.path.getsize(fp)
            self.send_response(200)
            self.send_header('Content-Type', 'application/octet-stream')
            self.send_header('Content-Disposition', f'attachment; filename="{name}"')
            self.send_header('Content-Length', str(size))
            self.end_headers()
            with open(fp, 'rb') as f:
                while True:
                    c = f.read(1 << 16)
                    if not c: break
                    self.wfile.write(c)
        else:
            self._json({'error': 'not found'}, 404)
    def do_POST(self):
        sweep()
        if self.path != '/api/backup': self._json({'error': 'not found'}, 404); return
        try:
            ln = int(self.headers.get('Content-Length', 0))
            d = json.loads(self.rfile.read(ln) or b'{}')
        except Exception:
            self._json({'error': 'invalid body'}, 400); return
        m = FOLDER_RE.search(d.get('url', ''))
        if not m: self._json({'error': '唔係有效嘅 Drive folder link'}, 400); return
        with _lock:
            active = len(_active)
            if active >= MAX_ACTIVE: self._json({'error': 'Server 繁忙，請稍後再試'}, 429); return
            job = uuid.uuid4().hex
            _active[job] = time.time()
        threading.Thread(target=run_job, args=(job, m.group(1)), daemon=True).start()
        self._json({'job': job})


def watch():
    while True:
        time.sleep(900); sweep()


if __name__ == '__main__':
    os.makedirs(BASE, exist_ok=True)
    threading.Thread(target=watch, daemon=True).start()
    print('listening :8301')
    ThreadingHTTPServer(('127.0.0.1', 8301), H).serve_forever()