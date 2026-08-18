#!/usr/bin/env python3
"""Sanity tests — 深測 parse_gem / list_folder / make_xlsx chunking / dedupe / path guard."""
import glob, io, json, os, re, sys, urllib.request

sys.path.insert(0, '/home/ubuntu/gembackup_web')
import app

PASS = 0
def ok(cond, msg):
    global PASS
    assert cond, 'FAIL: ' + msg
    PASS += 1
    print('  ok -', msg)

FIXTURES = '/home/ubuntu/gem_backup/raw'           # optional: local .bin fixtures
TEST_FOLDER = os.environ.get('TEST_FOLDER_ID', '') # optional: live folder to verify against

# ── 1. parse_gem vs real raw fixtures (skip if not present) ──
print('[1] parse_gem vs real .bin fixtures')
bins = glob.glob(FIXTURES + '/*.bin')
if bins:
    ok(len(bins) == 34, f'{len(bins)} raw bins exist')
    total_rows = 0
    for b in bins:
        rows = app.parse_gem(open(b, 'rb').read())
        ok(len(rows) >= 1, f'{os.path.basename(b)} parses >=1 row')
        total_rows += len(rows)
    ok(total_rows >= 34, f'total rows {total_rows} >= 34')
    sudoku = app.parse_gem(open(FIXTURES + '/1sUNg-5TkbIX3N3LWQrr01cSQBnhUv9bs.bin', 'rb').read())[0]
    ok(sudoku[0] == 'Sudoku', 'Sudoku name')
    ok(sudoku[2].startswith('You are an expert'), 'Sudoku instructions intact')
else:
    print('  skip (no fixtures at %s)' % FIXTURES)

# ── 2. list_folder vs live Drive folder (skip unless TEST_FOLDER_ID set) ──
print('[2] list_folder vs live Drive folder')
if TEST_FOLDER:
    entries = app.list_folder(TEST_FOLDER)
    ok(len(entries) >= 1, f'{len(entries)} entries found')
    names = {e.name for e in entries}
    ok(any('Sudoku' in n for n in names), 'known gem present')
    ok(all(e.kind in ('gem', 'file', 'folder') for e in entries), 'kinds valid')
else:
    print('  skip (set TEST_FOLDER_ID to verify against a live folder)')

# ── 3. make_xlsx 長 cell chunking + manifest sheet ──
print('[3] xlsx >32767 chunking + manifest sheet')
long_instr = 'A' * 40000
long_desc = 'B' * 35000
from openpyxl import load_workbook
tmp = '/tmp/sanity.xlsx'
app.make_xlsx([app.GemRow('LongGem', long_desc, long_instr, 'fid123', 'Aug 1')], tmp,
              {'Backup time': '2026-08-18T00:00:00+00:00', 'Tool version': app.VERSION,
               'Source folder name': 'My Gems', 'Source folder ID': 'fld1',
               'Gems found': 1, 'Attachments': 0, 'Total size (MB)': 0.1, 'Failed files': 0})
wb = load_workbook(tmp)
ok(wb.sheetnames == ['Manifest', 'Gems'], f'sheets = Manifest + Gems (got {wb.sheetnames})')
ws = wb['Gems']
ok(ws.max_row == 3, f'3 rows (1 + 2 continuations), got {ws.max_row}')
maxlen = 0; joined = ''
for r in ws.iter_rows(min_row=2, max_col=3):
    for c in r:
        v = c.value or ''
        maxlen = max(maxlen, len(v))
        if c.column in (2, 3): joined += v
ok(maxlen <= app.CELL_MAX, f'no cell > {app.CELL_MAX} (max {maxlen})')
desc_joined = ''.join((c.value or '') for r in ws.iter_rows(min_row=2) for c in r if c.column == 2)
instr_joined = ''.join((c.value or '').replace('【續上】', '') for r in ws.iter_rows(min_row=2) for c in r if c.column == 3)
ok(desc_joined == long_desc, f'desc lossless ({len(desc_joined)} chars)')
ok(instr_joined == long_instr, f'instr lossless ({len(instr_joined)} chars)')
ok(ws.cell(3, 3).value.startswith('【續上】'), 'continuation marker present')
ok(ws.cell(2, 4).value == 'https://gemini.google.com/gem/fid123', 'link only on first chunk')
ok(ws.cell(2, 5).value == 'fid123', 'Drive file ID column')
mws = wb['Manifest']
vals = {mws.cell(r, 1).value: mws.cell(r, 2).value for r in range(2, mws.max_row + 1)}
ok(vals.get('Tool version') == app.VERSION and vals.get('Source folder name') == 'My Gems'
   and vals.get('Gems found') == 1, 'manifest sheet fields')

# ── 4. dedupe_name ──
print('[4] zip name dedupe')
used = {}
a = app.dedupe_name('a_b.jpg', used)
b = app.dedupe_name('a_b.jpg', used)
c = app.dedupe_name('file.txt', used)
d = app.dedupe_name('noext', used)
e = app.dedupe_name('noext', used)
ok(a == 'a_b.jpg' and b == 'a_b (1).jpg', f'dup suffix ({a},{b})')
ok(c == 'file.txt', 'single ok')
ok(d == 'noext' and e == 'noext_1', f'no-ext dup ({d},{e})')
ok(len({a, b, c, d, e}) == 5, 'all distinct')

# ── 5. list_folder recursion: cycle guard (synthetic HTML) ──
print('[5] recursion + cycle guard')
def fake_get(url, timeout=30):
    h = ('<div class="flip-entry" id="entry-AAAA".*?'
         'href="https://drive.google.com/drive/folders/AAAA".*?'
         'flip-entry-title">self</div>.*?'
         'flip-entry-last-modified"><div>x</div>')
    return type('R', (), {'raise_for_status': lambda s: None, 'text': h})()
orig = app.requests.get
app.requests.get = fake_get
try:
    out = app.list_folder('AAAA')
    ok(out == [], f'cycle guard returns empty (got {len(out)})')
finally:
    app.requests.get = orig

def fake_get2(url, timeout=30):
    h = ('<div class="flip-entry" id="entry-SUB".*?'
         'href="https://drive.google.com/drive/folders/SUBFOLDER".*?'
         'flip-entry-title">sub</div>.*?'
         'flip-entry-last-modified"><div>x</div>'
         '<div class="flip-entry" id="entry-G1".*?'
         'href="https://gemini.google.com/gem/G1".*?'
         'flip-entry-title">Gem1</div>.*?'
         'flip-entry-last-modified"><div>x</div>')
    return type('R', (), {'raise_for_status': lambda s: None, 'text': h})()
app.requests.get = fake_get2
try:
    out = app.list_folder('ROOT')
    ok(any(e.kind == 'gem' and e.fid == 'G1' for e in out), 'nested gem found')
finally:
    app.requests.get = orig

# ── 6. path traversal guard + job_state validation ──
print('[6] security guards')
ok(app.job_state('../../etc/passwd') is None, 'traversal job id rejected')
ok(app.job_state('a' * 32) is None, 'nonexistent valid job -> None')
ok(app.job_state('') is None, 'empty job rejected')
ok(app.FOLDER_RE.search('https://drive.google.com/drive/u/0/folders/AbC123?usp=sharing').group(1) == 'AbC123', 'u/0 URL variant')
ok(app.FOLDER_RE.search('https://drive.google.com/drive/folders/xyz_123') is not None, 'bare folder URL')
ok(app.FOLDER_RE.search('https://evil.com/drive/folders/xyz') is None, 'non-google URL rejected')

# ── 7. build_package offline: zip tree + manifest.json + disk_usage ──
print('[7] build_package offline (zip tree + manifest)')
import shutil, zipfile
app.BASE = '/tmp/gembackup_test'
shutil.rmtree(app.BASE, ignore_errors=True); os.makedirs(app.BASE)
orig_fetch, orig_dl = app.fetch_folder, app.dl_to_file
app.fetch_folder = lambda fid: ('My Gems', [
    app.Entry('g1', 'Gem1', 'Jun 1', 'gem'),
    app.Entry('a1', 'doc.pdf', 'Jun 2', 'file')])
def fake_dl(fid, dest, cap=app.MAX_FILE):
    if fid == 'g1':
        open(dest, 'wb').write(b'\x12\x10\n\x04Gem1\x12\x02hi\x1a\x04body')
    else:
        open(dest, 'wb').write(b'fake')
    return 'ab' * 32
app.dl_to_file = fake_dl
try:
    os.makedirs(os.path.join(app.BASE, 'testjob'), exist_ok=True)
    app.build_package('testjob', 'fld123')
    jd = os.path.join(app.BASE, 'testjob')
    z = zipfile.ZipFile(os.path.join(jd, 'Gems_Backup.zip'))
    names = z.namelist()
    ok('gem-backup/manifest.json' in names, 'zip has manifest.json')
    ok('gem-backup/gems/gem-001/metadata.json' in names
       and 'gem-backup/gems/gem-001/instructions.txt' in names
       and 'gem-backup/gems/gem-001/description.txt' in names, 'zip gem tree')
    ok('gem-backup/attachments/doc.pdf' in names, 'zip attachment')
    man = json.loads(z.read('gem-backup/manifest.json'))
    ok(man['counts']['gems'] == 1 and man['counts']['attachments'] == 1, 'manifest json counts')
    ok(man['source_folder']['name'] == 'My Gems' and man['source_folder']['id'] == 'fld123', 'source folder in manifest')
    ok(man['tool'] == 'gem-backup' and man['version'] == app.VERSION, 'tool + version in manifest')
    ok(len(man['attachments'][0]['sha256']) == 64 and man['attachments'][0]['original_name'] == 'doc.pdf', 'sha256 + original name')
    ok(man['gems'][0]['drive_id'] == 'g1' and man['gems'][0]['status'] == 'ok', 'gem drive_id + status')
    ok(app.disk_usage() > 0, 'disk_usage counts job bytes')
finally:
    app.fetch_folder, app.dl_to_file = orig_fetch, orig_dl
    app.BASE = '/tmp/gembackup'
    shutil.rmtree('/tmp/gembackup_test', ignore_errors=True)

print(f'\nALL {PASS} CHECKS PASSED')